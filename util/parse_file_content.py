import os
from pathlib import Path

from docx2txt import process as docx2txt_process
# from ppt2txt import process as ppt2txt_process
from pptx2md import convert as pptx2md_parser,ConversionConfig


import openpyxl
import PyPDF2

def parse_txt_content(file_path):
    """
    解析TXT文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        print(f"解析TXT文件失败: {e}")
        return ""

def parse_docx_content(file_path):
    """
    解析Word (.docx) 文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    try:
        text = docx2txt_process(file_path)
        return text
    except Exception as e:
        print(f"解析DOCX文件失败: {e}")
        return ""

def parse_pptx_content(file_path):
    """
    解析PowerPoint (.pptx) 文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    try:
        text = pptx2md_parser(ConversionConfig(
            pptx_path=Path(file_path),
            output_path=Path(file_path.replace('.pptx', '.md')),
            image_dir=Path(file_path.replace('.pptx', '_images')),
            disable_notes=True
        ))
        with open(file_path.replace('.pptx', '.md'), 'r', encoding='utf-8') as f:
            text = f.read()

        # 删除临时生成的md文件
        os.remove(file_path.replace('.pptx', '.md'))
        # 删除临时生成的图片文件夹
        # 删除文件夹中所有文件
        for f in os.listdir(file_path.replace('.pptx', '_images')):
            os.remove(os.path.join(file_path.replace('.pptx', '_images'), f))
        os.rmdir(file_path.replace('.pptx', '_images'))

        return text
    except Exception as e:
        print(f"解析PPTX文件失败: {e}")
        return ""

def parse_excel_content(file_path):
    """
    解析Excel (.xlsx) 文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        full_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value is not None:
                        row_text.append(str(cell.value))
                if row_text:
                    full_text.append(' '.join(row_text))
        return '\n'.join(full_text)
    except Exception as e:
        print(f"解析Excel文件失败: {e}")
        return ""

def parse_pdf_content(file_path):
    """
    解析PDF文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    try:
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            full_text = []
            for page_num in range(len(reader.pages)):
                page = reader.pages[page_num]
                full_text.append(page.extract_text())
            return '\n'.join(full_text)
    except Exception as e:
        print(f"解析PDF文件失败: {e}")
        return ""

def upload_file(file_path):
    """
    根据文件类型解析文件内容
    :param file_path: 文件路径
    :return: 文件内容字符串
    """
    _, file_extension = os.path.splitext(file_path)
    file_extension = file_extension.lower()

    if file_extension == '.txt':
        return parse_txt_content(file_path)
    elif file_extension == '.docx':
        return parse_docx_content(file_path)
    elif file_extension == '.pptx':
        return parse_pptx_content(file_path)
    elif file_extension in ['.xlsx', '.xls']:
        return parse_excel_content(file_path)
    elif file_extension == '.pdf':
        return parse_pdf_content(file_path)
    else:
        print(f"不支持的文件类型: {file_extension}")
        # 对于不支持的文件类型，可以尝试以二进制读取并返回部分内容，或者直接返回空字符串
        # try:
        #     with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        #         # 读取前N个字符作为预览，避免读取整个大型未知文件
        #         return f.read(1024*1024*99) # 读取前1KB内容
        # except Exception as e:
        #     print(f"尝试读取未知文件类型失败: {e}")
        return ""


# 示例用法
if __name__ == '__main__':


    test_files = [
        # "20XX大三学生经典入党申请书-入党申请.docx",
        # "公文网站 待做.txt"
        "111.pptx",
        # "工作簿1.xlsx",
        # "商业计划书PDF版.pdf"

    ]

    for test_file in test_files:
        print(f"\n--- 解析文件: {test_file} ---")
        content = upload_file(os.path.join(r"E:\project\file_renaming\test", test_file))
        print(content)

